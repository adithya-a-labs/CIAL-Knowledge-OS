"""CSV, XLSX, and standalone HTML reports for one Phase 3 run."""

from __future__ import annotations

import csv
import html
import json
import logging
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def write_results_csv(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> Path:
    """Write the established UTF-8-with-BOM batch CSV schema."""

    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        writer.writerows(rows)
    logger.info(
        "csv_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _json_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        parsed = json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return []
    return parsed if isinstance(parsed, list) else []


def write_results_xlsx(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> Path:
    """Write a formatted workbook with clickable first-citation PDF links."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Install the pinned project "
            "dependencies from requirements.txt."
        ) from exc

    target = Path(path).expanduser().resolve()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Phase 3 Results"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for column_index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=name)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pdf_column = columns.index("pdf_links") + 1 if "pdf_links" in columns else None
    for row_index, row in enumerate(rows, start=2):
        for column_index, name in enumerate(columns, start=1):
            value = row.get(name, "")
            if isinstance(value, (dict, list, tuple)):
                value = json.dumps(value, ensure_ascii=False)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if pdf_column is not None:
            links = _json_list(row.get("pdf_links"))
            if links:
                cell = sheet.cell(row=row_index, column=pdf_column)
                cell.value = "Open first cited PDF"
                cell.hyperlink = str(links[0])
                cell.style = "Hyperlink"

    width_overrides = {
        "question": 42,
        "answer": 72,
        "retrieval_trace": 55,
        "error": 42,
        "pdf_links": 28,
    }
    for index, name in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            width_overrides.get(name, min(max(len(name) + 2, 14), 28))
        )
    sheet.row_dimensions[1].height = 28
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    logger.info(
        "xlsx_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _metric_card(label: str, value: Any) -> str:
    return (
        '<div class="metric"><span>'
        + html.escape(label)
        + "</span><strong>"
        + html.escape(str(value))
        + "</strong></div>"
    )


def write_latency_svg(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
) -> Path:
    """Write a dependency-free latency chart for the run figures directory."""

    values = [float(row.get("total_latency_seconds") or 0.0) for row in rows]
    width, height, margin = 900, 320, 45
    chart_width = width - 2 * margin
    chart_height = height - 2 * margin
    maximum = max(values, default=0.0) or 1.0
    bar_width = chart_width / max(len(values), 1)
    bars = []
    for index, value in enumerate(values):
        bar_height = chart_height * value / maximum
        x = margin + index * bar_width + 2
        y = margin + chart_height - bar_height
        bars.append(
            f'<rect x="{x:.2f}" y="{y:.2f}" width="{max(1.0, bar_width - 4):.2f}" '
            f'height="{bar_height:.2f}" fill="#1f6f8b"><title>Q{index + 1}: '
            f"{value:.4f}s</title></rect>"
        )
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}" role="img" aria-label="Question latency">'
        '<rect width="100%" height="100%" fill="#ffffff"/>'
        f'<line x1="{margin}" y1="{margin + chart_height}" x2="{width - margin}" '
        f'y2="{margin + chart_height}" stroke="#60717a"/>'
        f'<text x="{margin}" y="24" font-family="sans-serif" font-size="18" '
        f'fill="#12344d">Question latency (maximum {maximum:.4f}s)</text>'
        + "".join(bars)
        + "</svg>"
    )
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(svg, encoding="utf-8")
    logger.info(
        "latency_figure_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _citation_html(citation: Mapping[str, Any]) -> str:
    source = citation.get("source_file") or citation.get("source") or "Unknown"
    parts = [html.escape(str(source))]
    if citation.get("page_number") not in {None, ""}:
        parts.append(f"page {html.escape(str(citation['page_number']))}")
    if citation.get("chunk_id") not in {None, ""}:
        parts.append(f"chunk {html.escape(str(citation['chunk_id']))}")
    label = " | ".join(parts)
    link = citation.get("pdf_link")
    return (
        f'<li><a href="{html.escape(str(link), quote=True)}">{label}</a></li>'
        if link
        else f"<li>{label}</li>"
    )


def write_standalone_html(
    path: str | Path,
    *,
    rows: Sequence[Mapping[str, Any]],
    responses: Sequence[Mapping[str, Any] | None],
    summary: Mapping[str, Any],
    metrics: Mapping[str, Any],
    title: str = "CIAL Knowledge OS — Phase 3 Hybrid Retrieval",
) -> Path:
    """Write one offline report with embedded styles, data, and charts."""

    question_sections: list[str] = []
    latency_bars: list[str] = []
    max_latency = max(
        (float(row.get("total_latency_seconds") or 0.0) for row in rows),
        default=0.0,
    )
    for index, row in enumerate(rows, start=1):
        response = responses[index - 1] if index <= len(responses) else None
        response = response or {}
        citations = response.get("citations")
        citation_items = (
            [_citation_html(item) for item in citations if isinstance(item, Mapping)]
            if isinstance(citations, Sequence)
            and not isinstance(citations, (str, bytes))
            else []
        )
        retrieved = response.get("context_stages")
        retrieved = (
            retrieved.get("compressed", [])
            if isinstance(retrieved, Mapping)
            else []
        )
        context_blocks = "".join(
            "<details><summary>"
            + html.escape(
                f"{item.get('source') or 'Unknown'} — "
                f"{item.get('chunk_id') or 'chunk unknown'}"
            )
            + "</summary><pre>"
            + html.escape(str(item.get("text") or ""))
            + "</pre></details>"
            for item in retrieved
            if isinstance(item, Mapping)
        )
        usage = response.get("token_usage")
        usage = usage if isinstance(usage, Mapping) else {}
        question_sections.append(
            f"""<article>
<h3>{index}. {html.escape(str(row.get("question") or ""))}</h3>
<div class="status">{html.escape(str(row.get("answer_status") or row.get("status") or ""))}</div>
<h4>Answer</h4><pre>{html.escape(str(row.get("answer") or ""))}</pre>
<h4>Citations</h4><ul>{''.join(citation_items) or '<li>No citations</li>'}</ul>
<h4>Retrieved Context</h4>{context_blocks or '<p>No retained context.</p>'}
<div class="grid">
{_metric_card("Retrieved chunks", row.get("retrieved_chunks", 0))}
{_metric_card("Final sections", row.get("final_context_sections", 0))}
{_metric_card("Context usage", f"{usage.get('used', 0)} / {usage.get('budget', 0)} {usage.get('budget_type', '')}")}
{_metric_card("Latency (s)", row.get("total_latency_seconds", 0))}
</div>
</article>"""
        )
        latency = float(row.get("total_latency_seconds") or 0.0)
        width = 0 if max_latency == 0 else round(100 * latency / max_latency, 2)
        latency_bars.append(
            f'<div class="bar-row"><span>Q{index}</span><i style="width:{width}%"></i>'
            f"<b>{latency:.3f}s</b></div>"
        )

    cards = "".join(
        _metric_card(label, value)
        for label, value in (
            ("Questions", summary.get("question_count", len(rows))),
            ("Successful", summary.get("successful_questions", 0)),
            ("Answered", summary.get("answered_questions", 0)),
            ("Safe failures", summary.get("insufficient_evidence_questions", 0)),
            ("Average latency", summary.get("average_latency_seconds", 0)),
            ("Retrieval mode", summary.get("retrieval_mode", "")),
        )
    )
    embedded_data = json.dumps(
        {"summary": summary, "metrics": metrics},
        ensure_ascii=False,
        default=str,
    ).replace("</", "<\\/")
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(title)}</title>
<style>
:root{{--navy:#12344d;--blue:#1f6f8b;--ice:#eef6f8;--ink:#18252d;--muted:#62737d}}
*{{box-sizing:border-box}}body{{margin:0;background:#f4f7f8;color:var(--ink);
font:15px/1.55 system-ui,-apple-system,Segoe UI,sans-serif}}
header{{background:linear-gradient(120deg,var(--navy),var(--blue));color:white;padding:42px max(5vw,24px)}}
main{{max-width:1200px;margin:auto;padding:28px}}section,article{{background:white;border:1px solid #dce5e8;
border-radius:12px;padding:22px;margin:18px 0;box-shadow:0 2px 8px #1232}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px}}
.metric{{background:var(--ice);border-left:4px solid var(--blue);padding:12px;border-radius:7px}}
.metric span{{display:block;color:var(--muted);font-size:12px;text-transform:uppercase}}.metric strong{{font-size:20px}}
pre{{white-space:pre-wrap;overflow-wrap:anywhere;background:#f7f9fa;padding:14px;border-radius:7px}}
.status{{display:inline-block;background:#dbeef2;color:#16495b;padding:4px 9px;border-radius:999px}}
.bar-row{{display:grid;grid-template-columns:36px 1fr 72px;gap:8px;align-items:center;margin:8px 0}}
.bar-row i{{display:block;min-width:2px;height:18px;background:var(--blue);border-radius:3px}}
details{{border:1px solid #dce5e8;border-radius:6px;padding:8px;margin:8px 0}}a{{color:#0c607a}}
</style></head><body>
<header><h1>{html.escape(title)}</h1><p>Standalone, offline, evidence-aware run report.</p></header>
<main>
<section><h2>Executive Summary</h2><div class="grid">{cards}</div></section>
<section><h2>Metrics</h2><pre>{html.escape(json.dumps(metrics, indent=2, ensure_ascii=False, default=str))}</pre></section>
<section><h2>Charts</h2><h3>Question latency</h3>{''.join(latency_bars) or '<p>No latency data.</p>'}</section>
<section><h2>Questions, Answers, Evidence, Token Usage, and Retrieval Statistics</h2>
{''.join(question_sections) or '<p>No questions were processed.</p>'}</section>
<script type="application/json" id="run-data">{embedded_data}</script>
</main></body></html>"""
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(document, encoding="utf-8")
    logger.info(
        "html_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target
