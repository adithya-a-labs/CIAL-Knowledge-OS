from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cial_knowledge_os.batch_qa import CSV_COLUMNS, PHASE2_CSV_COLUMNS, PHASE3_CSV_COLUMNS
from cial_knowledge_os.config import Phase3Config
from cial_knowledge_os.phase3_runner import Phase3Runner
from cial_knowledge_os.run_manager import RunManager


class _ArtifactPipeline:
    def __init__(self, config: Phase3Config) -> None:
        self.config = config
        self.metrics = {
            "retrieval_latency": 0.01,
            "generation_latency": 0.02,
        }

    @property
    def is_ready_for_answering(self) -> bool:
        return True

    def answer(self, question: str):
        evidence = {
            "text": "Grounded evidence.",
            "score": 0.03,
            "rrf_score": 0.03,
            "retrieval_sources": ["dense", "bm25"],
            "source": "manual.pdf",
            "page_number": 4,
            "chunk_id": "manual:p4:c1",
            "metadata": {
                "source": str(self.config.project_root / "manual.pdf"),
                "file_name": "manual.pdf",
                "page_number": 4,
                "chunk_id": "manual:p4:c1",
                "chunk_index": 1,
            },
        }
        link = (self.config.project_root / "manual.pdf").resolve().as_uri() + "#page=4"
        return {
            "question": question,
            "answer": "Grounded answer [1].",
            "raw_answer": "Grounded answer [1].",
            "answer_status": "answered",
            "retrieval_mode": "hybrid",
            "retrieved": [evidence],
            "query_variants": [{"technique": "original", "query": question}],
            "retrieved_by_query": {"original": [evidence]},
            "context_stages": {
                "retrieved": [evidence],
                "deduplicated": [evidence],
                "expanded": [evidence],
                "merged": [evidence],
                "compressed": [evidence],
            },
            "stage_counts": {
                "retrieved": 1,
                "deduplicated": 1,
                "expanded": 1,
                "merged": 1,
                "compressed": 1,
            },
            "context": "Grounded evidence.",
            "prompt": "Grounded prompt.",
            "token_usage": {
                "budget": 100,
                "used": 20,
                "remaining": 80,
                "context_tokens": 20,
                "encoding_name": "cl100k_base",
                "truncated_sections": 0,
                "omitted_sections": 0,
                "budget_type": "tokens",
            },
            "citations": [
                {
                    "reference_id": 1,
                    "source": "manual.pdf",
                    "source_file": "manual.pdf",
                    "source_path": str(self.config.project_root / "manual.pdf"),
                    "page_number": 4,
                    "chunk_id": "manual:p4:c1",
                    "pdf_link": link,
                }
            ],
        }


class RunManagerTests(unittest.TestCase):
    def test_run_directories_never_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config = Phase3Config(project_root=Path(directory))
            first = RunManager.from_config(config).create()
            second = RunManager.from_config(config).create()
        self.assertNotEqual(first.root, second.root)


class Phase3ArtifactTests(unittest.TestCase):
    def test_complete_standalone_bundle_is_generated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_ArtifactPipeline(config),
                config=config,
            ).run(questions=["What is the control?"])
            paths = result.paths

            required_files = (
                paths.results_csv,
                paths.results_xlsx,
                paths.report_html,
                paths.config_json,
                paths.summary_json,
                paths.retrieval_json,
                paths.metrics_json,
                paths.logs,
            )
            self.assertTrue(all(path.is_file() for path in required_files))
            self.assertTrue(paths.figures.is_dir())
            self.assertTrue(
                (paths.figures / config.artifact_names.latency_figure).is_file()
            )
            self.assertEqual(len(list(paths.context.glob("*.md"))), 1)

            csv_header = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[0].split(",")
            self.assertEqual(
                csv_header,
                [*CSV_COLUMNS, *PHASE2_CSV_COLUMNS, *PHASE3_CSV_COLUMNS],
            )
            csv_values = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[1]
            self.assertIn("cl100k_base", csv_values)
            workbook = load_workbook(paths.results_xlsx)
            sheet = workbook.active
            pdf_column = csv_header.index("pdf_links") + 1
            self.assertIsNotNone(sheet.cell(2, pdf_column).hyperlink)

            report = paths.report_html.read_text(encoding="utf-8")
            self.assertIn("Executive Summary", report)
            self.assertIn("Retrieved Context", report)
            self.assertIn("Token Usage", report)
            self.assertNotIn("https://cdn", report)
            self.assertNotIn("<script src=", report)
            self.assertEqual(
                json.loads(paths.summary_json.read_text(encoding="utf-8"))[
                    "question_count"
                ],
                1,
            )
            self.assertEqual(
                json.loads(paths.config_json.read_text(encoding="utf-8"))[
                    "tokenizer_encoding_name"
                ],
                "cl100k_base",
            )
            self.assertIn(
                '"event":"run"',
                paths.logs.read_text(encoding="utf-8"),
            )

    def test_optional_run_metadata_is_exported_without_changing_default_schema(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_ArtifactPipeline(config),
                config=config,
            ).run(
                questions=["What is the control?"],
                run_metadata={
                    "run_type": "manual_qa",
                    "run_label": "interactive_manual_qa",
                },
            )

            csv_text = result.paths.results_csv.read_text(encoding="utf-8-sig")
            header = csv_text.splitlines()[0].split(",")
            self.assertEqual(header[-2:], ["run_type", "run_label"])
            self.assertIn("interactive_manual_qa", csv_text)

            config_payload = json.loads(
                result.paths.config_json.read_text(encoding="utf-8")
            )
            self.assertEqual(
                config_payload["run_overrides"]["run_type"],
                "manual_qa",
            )
            summary = json.loads(
                result.paths.summary_json.read_text(encoding="utf-8")
            )
            self.assertEqual(summary["run_label"], "interactive_manual_qa")
            retrieval = json.loads(
                result.paths.retrieval_json.read_text(encoding="utf-8")
            )
            self.assertEqual(
                retrieval[0]["run_metadata"]["run_type"],
                "manual_qa",
            )
            report = result.paths.report_html.read_text(encoding="utf-8")
            self.assertIn("Run Label", report)
            self.assertIn("interactive_manual_qa", report)


if __name__ == "__main__":
    unittest.main()
