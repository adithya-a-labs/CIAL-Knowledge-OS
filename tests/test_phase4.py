from __future__ import annotations

import json
import io
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from typing import Any
from unittest.mock import patch

from openpyxl import load_workbook

from cial_knowledge_os.batch_qa import (
    CSV_COLUMNS,
    PHASE2_CSV_COLUMNS,
    PHASE3_CSV_COLUMNS,
    PHASE4_CSV_COLUMNS,
)
from cial_knowledge_os.config import (
    KnowledgeOSConfig,
    Phase2Config,
    Phase3Config,
    Phase4Config,
)
from cial_knowledge_os.evidence_quality import EvidenceQualityScorer
from cial_knowledge_os.evidence_selector import EvidenceSelector
from cial_knowledge_os.phase4_pipeline import Phase4RAGPipeline
from cial_knowledge_os.phase4_runner import Phase4Runner
from cial_knowledge_os.phase4_trace import Phase4Trace
from cial_knowledge_os.reranker import CrossEncoderReranker, MockReranker
from cial_knowledge_os.token_budget import TokenBudgetManager


class _CharacterTokenizer:
    def encode(self, text: str, **_: Any) -> list[int]:
        return [ord(character) for character in text]

    def decode(self, values: list[int], **_: Any) -> str:
        return "".join(chr(value) for value in values)


class _StaticRetriever:
    def __init__(self, name: str, results: list[dict[str, Any]]) -> None:
        self.name = name
        self.results = results

    def retrieve(self, query: str, *, top_k: int) -> list[dict[str, Any]]:
        return [dict(item) for item in self.results[:top_k]]


class _CitingLLM:
    def invoke(self, prompt: str) -> str:
        self.prompt = prompt
        return "**Use the selected control** [1]."


class _FakeCrossEncoder:
    def __init__(self) -> None:
        self.calls: list[tuple[Any, ...]] = []

    def predict(self, pairs, **kwargs):
        self.calls.append((pairs, kwargs))
        return [0.2, 0.9]


class _ReadyPipeline:
    def __init__(self, pipeline: Phase4RAGPipeline) -> None:
        self.pipeline = pipeline
        self.config = pipeline.config
        self.metrics = pipeline.metrics
        self.token_manager = pipeline.token_manager

    @property
    def is_ready_for_answering(self) -> bool:
        return True

    def answer(self, question: str):
        return self.pipeline.answer(question)


def _candidate(
    chunk: str,
    text: str,
    *,
    source: str,
    page: int,
    score: float = 0.5,
) -> dict[str, Any]:
    return {
        "id": chunk,
        "text": text,
        "score": score,
        "source": Path(source).name,
        "page_number": page,
        "chunk_id": chunk,
        "metadata": {
            "source": source,
            "file_name": Path(source).name,
            "page_number": page,
            "chunk_id": chunk,
            "chunk_index": page,
        },
    }


class RerankerAndSelectionTests(unittest.TestCase):
    def test_cross_encoder_interface_preserves_phase3_scores(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker(
            "local-test-model",
            model=model,
            batch_size=2,
        )
        candidates = [
            _candidate("a", "weak", source="C:/docs/a.pdf", page=1, score=0.7),
            _candidate("b", "strong", source="C:/docs/b.pdf", page=2, score=0.4),
        ]

        result = reranker.rerank("Which control?", candidates)

        self.assertEqual([item["chunk_id"] for item in result.candidates], ["b", "a"])
        self.assertEqual(result.candidates[0]["reranker_score"], 0.9)
        self.assertEqual(result.candidates[0]["score"], 0.4)
        self.assertEqual(result.candidates[0]["original_rrf_rank"], 2)
        self.assertEqual(model.calls[0][1]["batch_size"], 2)

    def test_reranker_loads_from_cache_before_considering_download(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker("approved/reranker")
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            return_value=model,
        ) as cross_encoder, redirect_stdout(output):
            loaded = reranker._load_model()

        self.assertIs(loaded, model)
        self.assertEqual(reranker.load_source, "cache")
        self.assertTrue(cross_encoder.call_args.kwargs["local_files_only"])
        self.assertIn("loaded from local Hugging Face cache", output.getvalue())

    def test_reranker_downloads_once_after_cache_miss(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=False,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=[OSError("cache miss"), model],
        ) as cross_encoder, redirect_stdout(output):
            loaded = reranker._load_model()

        self.assertIs(loaded, model)
        self.assertEqual(reranker.load_source, "download")
        self.assertEqual(
            [
                call.kwargs["local_files_only"]
                for call in cross_encoder.call_args_list
            ],
            [True, False],
        )
        self.assertIn(
            'Downloading reranker model "approved/reranker"',
            output.getvalue(),
        )
        self.assertIn(
            "Reranker downloaded and cached successfully",
            output.getvalue(),
        )

    def test_enterprise_offline_mode_never_attempts_download(self) -> None:
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=True,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=OSError("cache miss"),
        ) as cross_encoder, redirect_stdout(output):
            with self.assertRaisesRegex(
                RuntimeError,
                'Configured reranker model: "approved/reranker"',
            ) as context:
                reranker._load_model()

        self.assertEqual(cross_encoder.call_count, 1)
        self.assertTrue(cross_encoder.call_args.kwargs["local_files_only"])
        self.assertIn("Local-only mode: enabled", str(context.exception))
        self.assertIn("MockReranker", str(context.exception))
        self.assertIn(
            "Download skipped because enterprise offline mode is enabled",
            output.getvalue(),
        )

    def test_download_failure_message_is_actionable(self) -> None:
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=False,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=[
                OSError("cache miss"),
                OSError("network unavailable"),
            ],
        ), redirect_stdout(output):
            with self.assertRaises(RuntimeError) as context:
                reranker._load_model()

        message = str(context.exception)
        self.assertIn('Configured reranker model: "approved/reranker"', message)
        self.assertIn("Local-only mode: disabled", message)
        self.assertIn("stage it manually", message)
        self.assertIn("MockReranker", message)

    def test_mock_reranker_is_deterministic(self) -> None:
        candidates = [
            _candidate("a", "alpha", source="C:/docs/a.pdf", page=1),
            _candidate("b", "beta", source="C:/docs/b.pdf", page=2),
        ]
        reranker = MockReranker({"a": 0.1, "b": 0.8})

        first = reranker.rerank("question", candidates)
        second = reranker.rerank("question", candidates)

        self.assertEqual(
            [item["chunk_id"] for item in first.candidates],
            [item["chunk_id"] for item in second.candidates],
        )
        self.assertEqual(first.candidates[0]["reranker_score"], 0.8)

    def test_selector_records_score_diversity_redundancy_and_budget_discards(
        self,
    ) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=100)
        selector = EvidenceSelector(
            manager,
            strategies=(
                "top_k",
                "reranker_score_threshold",
                "source_diversity",
                "redundancy_reduction",
                "token_budget",
            ),
            max_chunks=3,
            score_threshold=0.5,
            token_budget=28,
            max_chunks_per_source=1,
            redundancy_threshold=0.8,
        )
        candidates = [
            _candidate("a", "alpha control", source="C:/docs/a.pdf", page=1)
            | {"reranker_score": 0.95},
            _candidate("b", "alpha control", source="C:/docs/b.pdf", page=2)
            | {"reranker_score": 0.90},
            _candidate("c", "different control", source="C:/docs/a.pdf", page=3)
            | {"reranker_score": 0.85},
            _candidate("d", "weak", source="C:/docs/d.pdf", page=4)
            | {"reranker_score": 0.1},
            _candidate(
                "e",
                "this evidence is too long",
                source="C:/docs/e.pdf",
                page=5,
            )
            | {"reranker_score": 0.8},
        ]

        result = selector.select(candidates)

        self.assertEqual([item["chunk_id"] for item in result.selected], ["a"])
        reasons = {item["chunk_id"]: item["discard_reason"] for item in result.discarded}
        self.assertEqual(reasons["b"], "redundant")
        self.assertEqual(reasons["c"], "source_diversity")
        self.assertEqual(reasons["d"], "low_score")
        self.assertEqual(reasons["e"], "token_budget")
        self.assertLessEqual(result.selected_tokens, 28)

    def test_selector_respects_maximum_evidence_count(self) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=100)
        selector = EvidenceSelector(
            manager,
            strategies=("top_k",),
            max_chunks=2,
            score_threshold=0.0,
            token_budget=100,
            max_chunks_per_source=10,
            redundancy_threshold=1.0,
        )
        candidates = [
            _candidate(
                f"chunk-{index}",
                f"evidence {index}",
                source=f"C:/docs/{index}.pdf",
                page=index,
            )
            | {"reranker_score": 1.0 - index / 10}
            for index in range(1, 5)
        ]

        result = selector.select(candidates)

        self.assertEqual(len(result.selected), 2)
        self.assertTrue(
            all(
                item["discard_reason"] == "max_evidence_count"
                for item in result.discarded
            )
        )


class EvidenceQualityAndTraceTests(unittest.TestCase):
    def test_quality_scoring_reports_provenance_metadata_and_strength(self) -> None:
        scorer = EvidenceQualityScorer(
            strong_threshold=0.7,
            medium_threshold=0.4,
        )
        selected = [
            _candidate("a", "alpha", source="C:/docs/a.pdf", page=1)
            | {
                "reranker_score": 0.8,
                "retrieval_sources": ["dense", "bm25"],
                "evidence_token_count": 5,
            },
            _candidate("b", "beta", source="C:/docs/b.pdf", page=2)
            | {
                "reranker_score": 0.45,
                "retrieval_sources": ["dense"],
                "evidence_token_count": 4,
            },
        ]

        report = scorer.score(selected)

        self.assertEqual(report.chunks[0]["retrieval_source"], "both")
        self.assertEqual(report.chunks[0]["evidence_strength"], "strong")
        self.assertTrue(report.chunks[0]["metadata_complete"])
        self.assertEqual(report.summary["unique_source_count"], 2)
        self.assertEqual(
            report.summary["strength_distribution"],
            {"strong": 1, "medium": 1, "weak": 0},
        )

    def test_phase4_trace_round_trip(self) -> None:
        trace = Phase4Trace.from_dict(
            {
                "question": "What is the control?",
                "selected_chunks": [{"chunk_id": "a"}],
                "artifacts": {"context": Path("context/a.md")},
            }
        )

        restored = Phase4Trace.from_json(trace.to_json())

        self.assertEqual(restored.to_dict()["question"], "What is the control?")
        self.assertEqual(
            restored.to_dict()["artifacts"]["context"],
            str(Path("context/a.md")),
        )


class Phase4PipelineAndArtifactTests(unittest.TestCase):
    def _pipeline(self, root: Path) -> Phase4RAGPipeline:
        first = _candidate(
            "first",
            "The selected control requires review.",
            source=str(root / "manual.pdf"),
            page=1,
            score=0.8,
        )
        second = _candidate(
            "second",
            "Unrelated cafeteria information.",
            source=str(root / "other.pdf"),
            page=2,
            score=0.7,
        )
        config = Phase4Config(
            project_root=root,
            max_context_tokens=300,
            evidence_token_budget=120,
            max_query_variants=1,
            evidence_max_chunks=2,
            evidence_score_threshold=0.5,
            evidence_max_chunks_per_source=2,
        )
        return Phase4RAGPipeline(
            config,
            llm=_CitingLLM(),
            tokenizer=_CharacterTokenizer(),
            retrievers={
                "dense": _StaticRetriever("dense", [first, second]),
                "bm25": _StaticRetriever("bm25", [first, second]),
            },
            reranker=MockReranker({"first": 0.95, "second": 0.1}),
        )

    def test_pipeline_reranks_selects_and_preserves_phase3_keys(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            response = pipeline.answer("What control is required?")

        self.assertIn("context_stages", response)
        self.assertIn("token_usage", response)
        self.assertEqual(response["retrieval_mode"], "hybrid")
        self.assertEqual(
            [item["chunk_id"] for item in response["selected_evidence"]],
            ["first"],
        )
        self.assertEqual(
            response["discarded_evidence"][0]["discard_reason"],
            "low_score",
        )
        self.assertNotIn("cafeteria", response["context"])
        self.assertGreater(response["token_efficiency"]["token_reduction_percent"], 0)
        self.assertEqual(
            response["question_trace"]["pipeline_flow"][2],
            "reranker",
        )

    def test_pipeline_answer_downloads_once_then_uses_cached_model(self) -> None:
        cache_state = {"available": False}
        local_only_calls: list[bool] = []

        def model_factory(*_: Any, **kwargs: Any) -> _FakeCrossEncoder:
            local_only = bool(kwargs["local_files_only"])
            local_only_calls.append(local_only)
            if local_only and not cache_state["available"]:
                raise OSError("cache miss")
            if not local_only:
                cache_state["available"] = True
            return _FakeCrossEncoder()

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first_pipeline = self._pipeline(root)
            first_pipeline.reranker = CrossEncoderReranker(
                first_pipeline.config.reranker_model_name,
                local_files_only=False,
            )
            second_pipeline = self._pipeline(root)
            second_pipeline.reranker = CrossEncoderReranker(
                second_pipeline.config.reranker_model_name,
                local_files_only=False,
            )
            output = io.StringIO()
            with patch(
                "sentence_transformers.CrossEncoder",
                side_effect=model_factory,
            ), redirect_stdout(output):
                first_response = first_pipeline.answer("What control is required?")
                second_response = second_pipeline.answer("What control is required?")

        self.assertEqual(local_only_calls, [True, False, True])
        self.assertEqual(first_pipeline.reranker.load_source, "download")
        self.assertEqual(second_pipeline.reranker.load_source, "cache")
        self.assertTrue(first_response["answer"])
        self.assertTrue(second_response["answer"])
        self.assertIn("downloaded and cached successfully", output.getvalue())
        self.assertIn("loaded from local Hugging Face cache", output.getvalue())

    def test_phase4_bundle_contains_all_artifacts_and_report_sections(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            (root / "other.pdf").write_bytes(b"%PDF-test")
            pipeline = self._pipeline(root)
            result = Phase4Runner(
                pipeline=_ReadyPipeline(pipeline),
                config=pipeline.config,
            ).run(
                questions=["What control is required?"],
                run_mode="smoke",
            )

            paths = result.paths
            required = (
                paths.results_csv,
                paths.results_xlsx,
                paths.report_html,
                paths.config_json,
                paths.summary_json,
                paths.metrics_json,
                paths.retrieval_json,
                paths.logs,
            )
            self.assertTrue(all(path.is_file() for path in required))
            self.assertEqual(len(list(paths.context.glob("*.md"))), 1)
            self.assertGreaterEqual(len(list(paths.figures.glob("*.svg"))), 8)

            header = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[0].split(",")
            self.assertEqual(
                header[: len(CSV_COLUMNS + PHASE2_CSV_COLUMNS + PHASE3_CSV_COLUMNS)],
                CSV_COLUMNS + PHASE2_CSV_COLUMNS + PHASE3_CSV_COLUMNS,
            )
            for column in PHASE4_CSV_COLUMNS:
                self.assertIn(column, header)
            workbook = load_workbook(paths.results_xlsx)
            pdf_column = header.index("pdf_links") + 1
            self.assertIsNotNone(workbook.active.cell(2, pdf_column).hyperlink)

            report = paths.report_html.read_text(encoding="utf-8")
            for heading in (
                "Executive Summary",
                "Answers",
                "Citations",
                "Reranking Trace",
                "Evidence Selection",
                "Token Reduction",
                "Latency Breakdown",
                "Evidence Quality",
                "Source Diversity",
                "Selected vs Discarded Chunks",
                "Discard Reason Breakdown",
                "Phase 3 vs Phase 4 Comparison",
            ):
                self.assertIn(heading, report)
            self.assertNotIn("https://cdn", report)
            self.assertNotIn("<script src=", report)
            self.assertIn("<strong>Use the selected control</strong>", report)

            trace = json.loads(
                paths.retrieval_json.read_text(encoding="utf-8")
            )[0]
            self.assertEqual(trace["selected_chunks"][0]["chunk_id"], "first")
            self.assertEqual(
                trace["discarded_chunks"][0]["discard_reason"],
                "low_score",
            )
            self.assertIn("results_csv", trace["artifacts"])

    def test_previous_phase_defaults_remain_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            phase1 = KnowledgeOSConfig(project_root=root)
            phase2 = Phase2Config(project_root=root)
            phase3 = Phase3Config(project_root=root)
            phase4 = Phase4Config(project_root=root)

        self.assertEqual(phase1.qdrant_collection_name, "cial_basic_rag")
        self.assertEqual(phase2.qdrant_collection_name, "cial_phase2")
        self.assertEqual(phase3.qdrant_collection_name, "cial_phase3")
        self.assertEqual(phase3.phase_output_name, "03_Hybrid_Retrieval")
        self.assertEqual(phase4.phase_output_name, "04_Reranking_and_Evidence_Selection")
        self.assertFalse(phase4.enable_neighbor_expansion)
        self.assertFalse(phase4.reranker_local_files_only)


if __name__ == "__main__":
    unittest.main()
